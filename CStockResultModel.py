from CFnguideModel import  CFnguideData
from CStockCustomModel import  CStockCustomData
from openpyxl import load_workbook
import pandas as pd
import os

class CStockResultData:
    def __init__(self):
        self.__ResultDf = pd.DataFrame(
            columns=[
                self.__sTagStockCode,
                self.__sTagCompName,
                self.__sTagMarket,
                self.__sTagIndustry,
                self.__sTagLastQ,
                self.__sTagLastPrice,


                self.__sTagDomCapLastQ,
                self.__sTagDomCap_y1,
                self.__sTagDomCapPredQ,
                self.__sTagCapOrgLastQ,
                self.__sTagCapTotalLastQ,
                self.__sTagAssetLastQ,
                self.__sTagDebtLastQ,

                self.__sTagSales_y3,
                self.__sTagSales_y2,
                self.__sTagSales_y1,
                self.__sTagSalesLastQ,
                self.__sTagSalesLastQ_y1,
                self.__sTagSalesPredQ,

                self.__sTagBsProf_y3,
                self.__sTagBsProf_y2,
                self.__sTagBsProf_y1,
                self.__sTagBsProfLastQ,
                self.__sTagBsProfLastQ_y1,
                self.__sTagBsProfPredQ,

                self.__sTagNonBsProf_y1,
                self.__sTagNonBsProfLastQ,
                self.__sTagDomNetProf_y1,
                self.__sTagDomNetProfLastQ,
                self.__sTagDomNetProfLastQ_y1,
                self.__sTagDomNetProfPredQ,

                self.__sTagNetProfLastQ,
                self.__sTagBsProfBefTax_y1,
                self.__sTagBsProfBefTaxLastQ,

                self.__sTagSalesIncRate_y3_y2,
                self.__sTagSalesIncRate_y2_y1,
                self.__sTagSalesIncRate_lqy1_lq,
                self.__sTagSalesIncRateAvg,
                self.__sTagSalesIncRateTrend_y3_y1,
                self.__sTagSalesIncRateTrend_y3_PredQ,

                self.__sTagBsProfIncRate_y3_y2,
                self.__sTagBsProfIncRate_y2_y1,
                self.__sTagBsProfIncRate_lqy1_lq,
                self.__sTagBsProfIncRateAvg,
                self.__sTagBsProfIncRateTrend_y3_y1,
                self.__sTagBsProfIncRateTrend_y3_PredQ,

                self.__sTagCfBs_y1,
                self.__sTagCfBsLastQ,
                self.__sTagCfInv_y1,
                self.__sTagCfInvLastQ,
                self.__sTagCfFin_y1,
                self.__sTagCfFinLastQ,

                self.__sTagRoe_y1,
                self.__sTagRoe_y2,
                self.__sTagRoe_y3,
                self.__sTagRoePredQ,
                self.__sTagRoePredY,
                self.__sTagRoeConsen,
                # self.__sTagRoeBsProfPredQ,  # 영업이익으로 계산한 ROE
                # self.__sTagRoeBsProf_y1,  # 영업이익으로 계산한 ROE
                self.__sTagEps_y1,
                self.__sTagEps_y2,
                self.__sTagEps_y3,
                self.__sTagEpsPredQ,
                self.__sTagEpsPredY,
                self.__sTagEpsConsen,
                self.__sTagEpsIncrRate,
                # self.__sTagEpsBsProf_y1,
                # self.__sTagEpsBsProfPredQ,

                self.__sTagPer,
                self.__sTagPegr,
                self.__sTagIndusPer,
                self.__sTagPbr,
                self.__sTagDivIncome,

                self.__sTagSRim80Consen,
                self.__sTagSRim80PredQ,
                self.__sTagSRim80PredY,
                self.__sTagSRim80_y1,
                self.__sTagSRim90Consen,
                self.__sTagSRim90PredQ,
                self.__sTagSRim90PredY,
                self.__sTagSRim90_y1,
                self.__sTagSRim100Consen,
                self.__sTagSRim100PredQ,
                self.__sTagSRim100PredY,
                self.__sTagSRim100_y1,
                self.__sTagSRimPrice80Consen,
                self.__sTagSRimPrice80PredQ,
                self.__sTagSRimPrice80PredY,
                self.__sTagSRimPrice80_y1,
                self.__sTagSRimPrice90Consen,
                self.__sTagSRimPrice90PredQ,
                self.__sTagSRimPrice90PredY,
                self.__sTagSRimPrice90_y1,
                self.__sTagSRimPrice100Consen,
                self.__sTagSRimPrice100PredQ,
                self.__sTagSRimPrice100PredY,
                self.__sTagSRimPrice100_y1,
                self.__sTagSRimExpRate80Consen,
                self.__sTagSRimExpRate80PredQ,
                self.__sTagSRimExpRate80PredY,
                self.__sTagSRimExpRateAvg80,
                self.__sTagSRimExpRate80_y1,
                self.__sTagSRimExpRate90Consen,
                self.__sTagSRimExpRate90PredQ,
                self.__sTagSRimExpRate90PredY,
                self.__sTagSRimExpRateAvg90,
                self.__sTagSRimExpRate90_y1,
                self.__sTagSRimExpRate100Consen,
                self.__sTagSRimExpRate100PredQ,
                self.__sTagSRimExpRate100PredY,
                self.__sTagSRimExpRateAvg100,
                self.__sTagSRimExpRate100_y1,

                self.__sTagMarketCap,
                self.__sTagStockCntOrd,
                self.__sTagStockCntPref,
                self.__sTagStockCntTot,

                self.__sTagNetLoanRate,
                self.__sTagInterCovRate,

                self.__sTagNonBsRate_y1,
                self.__sTagNonBsRateLastQ,

                self.__sTagCfBsProfDiff_y1,
                self.__sTagCfBsProfDiffLastQ,
                self.__sTagCfPattern_y1,
                self.__sTagCfPatternLastQ,

                self.__sTagCortaxCond,
                # self.__sTagKPriceBsProf_y1,
                # self.__sTagKPriceBsProfPredQ,
                self.__sTagKPriceNetProf_y1,
                self.__sTagKPriceNetProfPredQ,
                self.__sTagKPriceNetProfPredY,
                self.__sTagKPriceNetProfConsen,
                # self.__sTagKPriceBsProfExpRate_y1,
                # self.__sTagKPriceBsProfExpRatePredQ,
                self.__sTagKPriceNetProfExpRate_y1,
                self.__sTagKPriceNetProfExpRatePredQ,
                self.__sTagKPriceNetProfExpRatePredY,
                self.__sTagKPriceNetProfExpRateConsen,

                self.__sTagPassFail,
                self.__sTagFailReason,
                self.__sBondBBB_5Rate,

            ])


    def AppendData(self, stockCode : str, fnData: CFnguideData , customData: CStockCustomData):
        dictData = {
            self.__sTagStockCode : stockCode,
            self.__sTagCompName : fnData.CompName(),
            self.__sTagMarket : fnData.Market(),
            self.__sTagIndustry : fnData.Industry(),
            self.__sTagLastQ : fnData.LastQ(),
            self.__sBondBBB_5Rate : customData.BondBBB_5Rate(),
            self.__sTagRoe_y1 : fnData.Roe_y1(),
            self.__sTagRoe_y2 : fnData.Roe_y2(),
            self.__sTagRoe_y3 : fnData.Roe_y3(),
            self.__sTagRoePredQ : customData.RoePredQ(),
            self.__sTagRoePredY : customData.RoePredY(),
            self.__sTagRoeConsen : fnData.RoeConsen(),
            # self.__sTagRoeBsProfPredQ : customData.RoeBsProfPredQ(),
            # self.__sTagRoeBsProf_y1 :customData.RoeBsProf_y1(),
            self.__sTagEps_y1 : fnData.Eps_y1(),
            self.__sTagEps_y2 : fnData.Eps_y2(),
            self.__sTagEps_y3 : fnData.Eps_y3(),
            self.__sTagEpsPredQ : customData.EpsPredQ(),
            self.__sTagEpsPredY: customData.EpsPredY(),
            self.__sTagEpsConsen : fnData.EpsConsen(),
            self.__sTagEpsIncrRate : fnData.EpsIncrRate(),
            # self.__sTagEpsBsProf_y1 : customData.EpsBsProf_y1(),
            # self.__sTagEpsBsProfPredQ : customData.EpsBsProfPredQ(),
            self.__sTagPer :fnData.Per(),
            self.__sTagIndusPer : fnData.IndusPer(),
            self.__sTagPbr : fnData.Pbr(),
            self.__sTagDivIncome : fnData.DivIncome(),
            self.__sTagCfBs_y1 : fnData.CfBs_y1(),
            self.__sTagCfBsLastQ : fnData.CfBsLastQ(),
            self.__sTagCfInv_y1 : fnData.CfInv_y1(),
            self.__sTagCfInvLastQ : fnData.CfInvLastQ(),
            self.__sTagCfFin_y1 : fnData.CfFin_y1(),
            self.__sTagCfFinLastQ : fnData.CfFinLastQ(),
            self.__sTagNonBsProf_y1 : customData.NonBsProf_y1(),
            self.__sTagNonBsProfLastQ : customData.NonBsProfLastQ(),
            self.__sTagSales_y3: fnData.Sales_y3(),
            self.__sTagSales_y2: fnData.Sales_y2(),
            self.__sTagSales_y1: fnData.Sales_y1(),
            self.__sTagSalesLastQ: fnData.SalesLastQ(),
            self.__sTagSalesLastQ_y1: fnData.SalesLastQ_y1(),
            self.__sTagSalesPredQ: customData.SalesPredQ(),
            self.__sTagBsProf_y3 : fnData.BsProf_y3(),
            self.__sTagBsProf_y2 : fnData.BsProf_y2(),
            self.__sTagBsProf_y1 : fnData.BsProf_y1(),
            self.__sTagBsProfLastQ : fnData.BsProfLastQ(),
            self.__sTagBsProfLastQ_y1 : fnData.BsProfLastQ_y1(),
            self.__sTagBsProfPredQ : customData.BsProfPredQ(),
            self.__sTagDomNetProf_y1 : fnData.DomNetProf_y1(),
            self.__sTagDomNetProfLastQ : fnData.DomNetProfLastQ(),
            self.__sTagDomNetProfLastQ_y1 : fnData.DomNetProfLastQ_y1(),
            self.__sTagDomNetProfPredQ : customData.DomNetProfPredQ(),
            self.__sTagNetProfLastQ : fnData.NetProfLastQ(),
            self.__sTagBsProfBefTax_y1 : fnData.BsProfBefTax_y1(),
            self.__sTagBsProfBefTaxLastQ : fnData.BsProfBefTaxLastQ(),
            self.__sTagSRim80Consen : customData.SRim80Consen(),
            self.__sTagSRim80PredQ : customData.SRim80PredQ(),
            self.__sTagSRim80PredY : customData.SRim80PredY(),
            self.__sTagSRim80_y1 : customData.SRim80_y1(),
            self.__sTagSRim90Consen : customData.SRim90Consen(),
            self.__sTagSRim90PredQ : customData.SRim90PredQ(),
            self.__sTagSRim90PredY : customData.SRim90PredY(),
            self.__sTagSRim90_y1 : customData.SRim90_y1(),
            self.__sTagSRim100Consen : customData.SRim100Consen(),
            self.__sTagSRim100PredQ : customData.SRim100PredQ(),
            self.__sTagSRim100PredY : customData.SRim100PredY(),
            self.__sTagSRim100_y1 : customData.SRim100_y1(),
            self.__sTagSRimPrice80Consen : customData.SRimPrice80Consen(),
            self.__sTagSRimPrice80PredQ : customData.SRimPrice80PredQ(),
            self.__sTagSRimPrice80PredY : customData.SRimPrice80PredY(),
            self.__sTagSRimExpRateAvg80 : customData.SRimExpRateAvg80(),
            self.__sTagSRimPrice80_y1 : customData.SRimPrice80_y1(),
            self.__sTagSRimPrice90Consen : customData.SRimPrice90Consen(),
            self.__sTagSRimPrice90PredQ : customData.SRimPrice90PredQ(),
            self.__sTagSRimPrice90PredY : customData.SRimPrice90PredY(),
            self.__sTagSRimExpRateAvg90 : customData.SRimExpRateAvg90(),
            self.__sTagSRimPrice90_y1 : customData.SRimPrice90_y1(),
            self.__sTagSRimPrice100Consen : customData.SRimPrice100Consen(),
            self.__sTagSRimPrice100PredQ : customData.SRimPrice100PredQ(),
            self.__sTagSRimPrice100PredY : customData.SRimPrice100PredY(),
            self.__sTagSRimExpRateAvg100 : customData.SRimExpRateAvg100(),
            self.__sTagSRimPrice100_y1 : customData.SRimPrice100_y1(),
            self.__sTagSRimExpRate80Consen : customData.SRimExpRate80Consen(),
            self.__sTagSRimExpRate80PredQ : customData.SRimExpRate80PredQ(),
            self.__sTagSRimExpRate80PredY : customData.SRimExpRate80PredY(),
            self.__sTagSRimExpRate80_y1 : customData.SRimExpRate80_y1(),
            self.__sTagSRimExpRate90Consen : customData.SRimExpRate90Consen(),
            self.__sTagSRimExpRate90PredQ : customData.SRimExpRate90PredQ(),
            self.__sTagSRimExpRate90PredY : customData.SRimExpRate90PredY(),
            self.__sTagSRimExpRate90_y1 : customData.SRimExpRate90_y1(),
            self.__sTagSRimExpRate100Consen : customData.SRimExpRate100Consen(),
            self.__sTagSRimExpRate100PredQ : customData.SRimExpRate100PredQ(),
            self.__sTagSRimExpRate100PredY : customData.SRimExpRate100PredY(),
            self.__sTagSRimExpRate100_y1 : customData.SRimExpRate100_y1(),
            self.__sTagDomCapLastQ : fnData.DomCapLastQ(),
            self.__sTagDomCap_y1 : fnData.DomCap_y1(),
            self.__sTagDomCapPredQ : customData.DomCapPredQ(),
            self.__sTagCapOrgLastQ : fnData.CapOrgLastQ(),
            self.__sTagCapTotalLastQ : fnData.CapTotalLastQ(),
            self.__sTagAssetLastQ : fnData.AssetLastQ(),
            self.__sTagDebtLastQ : fnData.DebtLastQ(),
            self.__sTagLastPrice : fnData.LastPrice(),
            self.__sTagMarketCap : fnData.MarketCap(),
            self.__sTagStockCntOrd : fnData.StockCntOrd(),
            self.__sTagStockCntPref : fnData.StockCntPref(),
            self.__sTagStockCntTot : fnData.StockCntTot(),
            self.__sTagNetLoanRate : fnData.NetLoanRate(),
            self.__sTagInterCovRate : fnData.InterCovRate(),
            self.__sTagNonBsRate_y1 : customData.NonBsRate_y1(),
            self.__sTagNonBsRateLastQ : customData.NonBsRateLastQ(),
            self.__sTagCfBsProfDiff_y1 : customData.CfBsProfDiff_y1(),
            self.__sTagCfBsProfDiffLastQ : customData.CfBsProfDiffLastQ(),
            self.__sTagCfPattern_y1 : customData.CfPattern_y1(),
            self.__sTagCfPatternLastQ : customData.CfPatternLastQ(),
            self.__sTagPegr : customData.Pegr(),
            self.__sTagCortaxCond : customData.CortaxCond(),
            # self.__sTagKPriceBsProf_y1 : customData.KPriceBsProf_y1(),
            # self.__sTagKPriceBsProfPredQ : customData.KPriceBsProfPredQ(),
            self.__sTagKPriceNetProf_y1 : customData.KPriceNetProf_y1(),
            self.__sTagKPriceNetProfPredQ : customData.KPriceNetProfPredQ(),
            self.__sTagKPriceNetProfPredY : customData.KPriceNetProfPredY(),
            self.__sTagKPriceNetProfConsen : customData.KPriceNetProfConsen(),
            # self.__sTagKPriceBsProfExpRate_y1 : customData.KPriceBsProfExpRate_y1(),
            # self.__sTagKPriceBsProfExpRatePredQ : customData.KPriceBsProfExpRatePredQ(),
            self.__sTagKPriceNetProfExpRate_y1 : customData.KPriceNetProfExpRate_y1(),
            self.__sTagKPriceNetProfExpRatePredQ : customData.KPriceNetProfExpRatePredQ(),
            self.__sTagKPriceNetProfExpRatePredY : customData.KPriceNetProfExpRatePredY(),
            self.__sTagKPriceNetProfExpRateConsen : customData.KPriceNetProfExpRateConsen(),
            self.__sTagPassFail : customData.PassFail(),
            self.__sTagFailReason : customData.FailReason(),

            self.__sTagSalesIncRate_y3_y2 : customData.SalesIncRate_y3_y2(),
            self.__sTagSalesIncRate_y2_y1 : customData.SalesIncRate_y2_y1(),
            self.__sTagSalesIncRate_lqy1_lq: customData.SalesIncRate_lqy1_lq(),
            self.__sTagSalesIncRateAvg: customData.SalesIncRateAvg(),
            self.__sTagSalesIncRateTrend_y3_y1: customData.SalesIncRateTrend_y3_y1(),
            self.__sTagSalesIncRateTrend_y3_PredQ: customData.SalesIncRateTrend_y3_PredQ(),

            self.__sTagBsProfIncRate_y3_y2: customData.BsProfIncRate_y3_y2(),
            self.__sTagBsProfIncRate_y2_y1: customData.BsProfIncRate_y2_y1(),
            self.__sTagBsProfIncRate_lqy1_lq: customData.BsProfIncRate_lqy1_lq(),
            self.__sTagBsProfIncRateAvg: customData.BsProfIncRateAvg(),
            self.__sTagBsProfIncRateTrend_y3_y1: customData.BsProfIncRateTrend_y3_y1(),
            self.__sTagBsProfIncRateTrend_y3_PredQ: customData.BsProfIncRateTrend_y3_PredQ()
        }

        self.__ResultDf = self.__ResultDf.append(dictData, ignore_index=True)

    def ClearData(self):
        self.__ResultDf.cle

    def ResultDf(self):
        return  self.__ResultDf

    def SummaryCols(self):
        summaryCols = [
            self.__sTagStockCode,
            self.__sTagCompName,
            self.__sTagIndustry,
            self.__sTagLastQ,
            self.__sTagPassFail,
            self.__sTagFailReason,
            self.__sTagLastPrice,
            self.__sTagMarketCap,

            self.__sTagSRimExpRate80Consen,
            self.__sTagSRimExpRate80PredQ,
            self.__sTagSRimExpRate80PredY,
            self.__sTagSRimExpRate80_y1,
            self.__sTagSRimExpRateAvg80,

            self.__sTagSRimPrice80Consen,
            self.__sTagSRimPrice80PredQ,
            self.__sTagSRimPrice80PredY,
            self.__sTagSRimPrice80_y1,

            self.__sTagSalesIncRateAvg,
            self.__sTagSalesIncRate_lqy1_lq,
            self.__sTagSalesIncRateTrend_y3_y1,
            self.__sTagSalesIncRateTrend_y3_PredQ,

            self.__sTagBsProfIncRateAvg,
            self.__sTagBsProfIncRate_lqy1_lq,
            self.__sTagBsProfIncRateTrend_y3_y1,
            self.__sTagBsProfIncRateTrend_y3_PredQ,

            self.__sTagNonBsRateLastQ,
            self.__sTagNonBsRate_y1,
            self.__sTagCfBsProfDiffLastQ,
            self.__sTagCfBsProfDiff_y1,

            self.__sTagCfPatternLastQ,
            self.__sTagCfPattern_y1,
            self.__sTagPegr,
            self.__sTagPer,
            self.__sTagIndusPer,
            self.__sTagPbr,
            self.__sTagDivIncome,

            self.__sTagNetLoanRate,
            self.__sTagInterCovRate,
        ]

        return summaryCols


    def SummaryGroupCols(self):
        groupCols = [
            self.__sTagStockGrp,
            self.__sTagStockGrp,
            self.__sTagStockGrp,
            self.__sTagStockGrp,
            self.__sTagStockGrp,
            self.__sTagStockGrp,
            self.__sTagStockGrp,
            self.__sTagStockGrp,

            self.__sTagSrimExpGrp,
            self.__sTagSrimExpGrp,
            self.__sTagSrimExpGrp,
            self.__sTagSrimExpGrp,
            self.__sTagSrimExpGrp,

            self.__sTagSrimPriceGrp,
            self.__sTagSrimPriceGrp,
            self.__sTagSrimPriceGrp,
            self.__sTagSrimPriceGrp,

            self.__sTagGrowthGrp,
            self.__sTagGrowthGrp,
            self.__sTagGrowthGrp,
            self.__sTagGrowthGrp,
            self.__sTagGrowthGrp,
            self.__sTagGrowthGrp,
            self.__sTagGrowthGrp,
            self.__sTagGrowthGrp,

            self.__sTagSafeIndGrp,
            self.__sTagSafeIndGrp,
            self.__sTagSafeIndGrp,
            self.__sTagSafeIndGrp,

            self.__sTagSubIndGrp,
            self.__sTagSubIndGrp,
            self.__sTagSubIndGrp,
            self.__sTagSubIndGrp,
            self.__sTagSubIndGrp,
            self.__sTagSubIndGrp,
            self.__sTagSubIndGrp,

            self.__sTagDebGrp,
            self.__sTagDebGrp
        ]

        resCols = [groupCols, self.SummaryCols()]

        return resCols

    # Column Name
    __sTagStockCode = '종목코드'
    __sTagCompName = '회사명'
    __sTagMarket = '시장'
    __sTagIndustry = '업종'
    __sTagLastQ = '회계분기'
    __sBondBBB_5Rate = '채권_EXP(%)'
    __sTagRoe_y1 = 'ROE[-Y1](%)'
    __sTagRoe_y2 = 'ROE[-Y2](%)'
    __sTagRoe_y3 = 'ROE[-Y3](%)'
    __sTagRoeConsen = 'ROE[C](%)'
    __sTagRoePredQ = 'ROE[Q+](%)'
    __sTagRoePredY = 'ROE[Y+](%)'
    # __sTagRoeBsProfPredQ = 'ROE_영업이익[-Y1](%)'
    # __sTagRoeBsProf_y1 = 'ROE_영업이익[Q+](%)'
    __sTagEps_y1 = 'EPS[-Y1]'
    __sTagEps_y2 = 'EPS[-Y2]'
    __sTagEps_y3 = 'EPS[-Y3]'
    __sTagEpsPredQ = 'EPS[Q+]'
    __sTagEpsPredY = 'EPS[Y+]'
    __sTagEpsConsen = 'EPS[C]'
    __sTagEpsIncrRate = 'EPS증가율(%)'
    # __sTagEpsBsProf_y1 = 'EPS_영업이익[-Y1]'
    # __sTagEpsBsProfPredQ = 'EPS_영업이익[Q+]'
    __sTagPer = 'PER[CUR]'
    __sTagIndusPer = '업종PER[CUR]'
    __sTagPbr = 'PBR[CUR]'
    __sTagDivIncome ='배당수익률[CUR](%)'
    __sTagCfBs_y1 = '영업CF[-Y1](억)'
    __sTagCfBsLastQ = '영업CF[Q](억)'
    __sTagCfInv_y1 = '투자CF[-Y1](억)'
    __sTagCfInvLastQ = '투자CF[Q](억)'
    __sTagCfFin_y1 = '재무CF[-Y1](억)'
    __sTagCfFinLastQ = '재무CF[Q](억)'
    __sTagNonBsProf_y1 = '비영업이익[-Y1](억)'
    __sTagNonBsProfLastQ = '비영업이익[Q](억)'
    __sTagDomNetProf_y1 = '지배순이익[-Y1](억)'
    __sTagDomNetProfLastQ = '지배순이익[Q](억)'
    __sTagDomNetProfLastQ_y1 = '지배순이익[Q-Y1](억)'
    __sTagDomNetProfPredQ = '지배순이익[Q+](억)'
    __sTagNetProfLastQ = '당기순이익[Q](억)'
    __sTagBsProfBefTax_y1 = '세전계속사업이익[-Y1](억)'
    __sTagBsProfBefTaxLastQ = '세전계속사업이익[Q](억)'
    __sTagSRim80Consen = 'SRIM_W80[C](억)'
    __sTagSRim80PredQ = 'SRIM_W80[Q+](억)'
    __sTagSRim80PredY = 'SRIM_W80[Y+](억)'
    __sTagSRim80_y1 = 'SRIM_W80[-Y1](억)'
    __sTagSRim90Consen = 'SRIM_W90[C](억)'
    __sTagSRim90PredQ = 'SRIM_W90[Q+](억)'
    __sTagSRim90PredY = 'SRIM_W90[Y+](억)'
    __sTagSRim90_y1 = 'SRIM_W90[-Y1](억)'
    __sTagSRim100Consen = 'SRIM_W100[C](억)'
    __sTagSRim100PredQ = 'SRIM_W100[Q+](억)'
    __sTagSRim100PredY = 'SRIM_W100[Y+](억)'
    __sTagSRim100_y1 = 'SRIM_W100[-Y1](억)'
    __sTagSRimPrice80Consen = '적정주가_W80[C]'
    __sTagSRimPrice80PredQ = '적정주가_W80[Q+]'
    __sTagSRimPrice80PredY = '적정주가_W80[Y+]'
    __sTagSRimPrice80_y1 = '적정주가_W80[-Y1]'
    __sTagSRimPrice90Consen = '적정주가_W90[C]'
    __sTagSRimPrice90PredQ = '적정주가_W90[Q+]'
    __sTagSRimPrice90PredY = '적정주가_W90[Y+]'
    __sTagSRimPrice90_y1 = '적정주가_W90[-Y1]'
    __sTagSRimPrice100Consen = '적정주가_W100[C]'
    __sTagSRimPrice100PredQ = '적정주가_W100[Q+]'
    __sTagSRimPrice100PredY = '적정주가_W100[Y+]'
    __sTagSRimPrice100_y1 = '적정주가_W100[-Y1]'
    __sTagSRimExpRate80Consen = 'SRIM_EXP_W80[C]'
    __sTagSRimExpRate80PredQ = 'SRIM_EXP_W80[Q+]'
    __sTagSRimExpRate80PredY = 'SRIM_EXP_W80[Y+]'
    __sTagSRimExpRate80_y1 = 'SRIM_EXP_W80[-Y1]'
    __sTagSRimExpRateAvg80 = 'SRIM_EXP_AVG_W80'
    __sTagSRimExpRate90Consen = 'SRIM_EXP_W90[C]'
    __sTagSRimExpRate90PredQ = 'SRIM_EXP_W90[Q+]'
    __sTagSRimExpRate90PredY = 'SRIM_EXP_W90[Y+]'
    __sTagSRimExpRate90_y1 = 'SRIM_EXP_W90[-Y1]'
    __sTagSRimExpRateAvg90 = 'SRIM_EXP_AVG_W90'
    __sTagSRimExpRate100Consen = 'SRIM_EXP_W100[C]'
    __sTagSRimExpRate100PredQ = 'SRIM_EXP_W100[Q+]'
    __sTagSRimExpRate100PredY = 'SRIM_EXP_W100[Y+]'
    __sTagSRimExpRate100_y1 = 'SRIM_EXP_W100[-Y1]'
    __sTagSRimExpRateAvg100 = 'SRIM_EXP_AVG_W100'
    __sTagDomCapLastQ = '지배주주자본[Q](억)'
    __sTagDomCap_y1 = '지배주주자본[-Y1](억)'
    __sTagDomCapPredQ = '지배주주자본[Q+](억)'
    __sTagCapOrgLastQ = '자본금[Q](억)'
    __sTagCapTotalLastQ = '자본[Q](억)'
    __sTagAssetLastQ = '자산[Q](억)'
    __sTagDebtLastQ = '부채[Q](억)'
    __sTagLastPrice = '최근종가'
    __sTagMarketCap = '시가총액(억)'
    __sTagStockCntOrd = '발행주식수_보통주'
    __sTagStockCntPref = '발행주식수_우선주'
    __sTagStockCntTot = '발행주식수_전체'
    __sTagNetLoanRate = '순차입금비율[Q](%)(<30)'
    __sTagInterCovRate = '이자보상배율[Q](>1)'
    __sTagNonBsRate_y1 = '비영업/영업이익[-Y1](%)(<30)'
    __sTagNonBsRateLastQ = '비영업/영업이익[Q](%)(<30)'
    __sTagCfBsProfDiff_y1 = '영업CF-영업이익[-Y1](>0)'
    __sTagCfBsProfDiffLastQ = '영업CF-영업이익[Q](>0)'
    __sTagCfPattern_y1 = '현금흐름패턴[-Y1](+--/-+-)'
    __sTagCfPatternLastQ = '현금흐름패턴[Q](+--/-+-)'
    __sTagPegr = 'PEGR[CUR](0.5/1.5)'
    __sTagCortaxCond = '자본잠식[Q]'
    # __sTagKPriceBsProf_y1 = 'Kprice_영업이익[-Y1]'
    # __sTagKPriceBsProfPredQ = 'Kprice_영업이익[Q+]'
    __sTagKPriceNetProf_y1 = 'Kprice_순이익[-Y1]'
    __sTagKPriceNetProfPredQ = 'Kprice_순이익[Q+]'
    __sTagKPriceNetProfPredY = 'Kprice_순이익[Y+]'
    __sTagKPriceNetProfConsen = 'Kprice_순이익[C]'
    # __sTagKPriceBsProfExpRate_y1 = 'Kprice_EXP_영업이익[-Y1]'
    # __sTagKPriceBsProfExpRatePredQ = 'Kprice_EXP_영업이익[Q+]'
    __sTagKPriceNetProfExpRate_y1 = 'Kprice_EXP_순이익[-Y1]'
    __sTagKPriceNetProfExpRatePredQ = 'Kprice_EXP_순이익[Q+]'
    __sTagKPriceNetProfExpRatePredY = 'Kprice_EXP_순이익[Y+]'
    __sTagKPriceNetProfExpRateConsen = 'Kprice_EXP_순이익[C]'

    __sTagPassFail = 'PASS/FAIL'
    __sTagFailReason = 'FAIL사유'

    # Sales 성장지표
    __sTagSales_y3 = '매출[-Y3](억)'
    __sTagSales_y2 = '매출[-Y2](억)'
    __sTagSales_y1 = '매출[-Y1](억)'
    __sTagSalesLastQ = '매출[Q](억)'
    __sTagSalesLastQ_y1 = '매출[Q-Y1](억)'
    __sTagSalesPredQ = '매출[Q+](억)'

    __sTagSalesIncRate_y3_y2 ='매출증가율[-Y3][-Y2](%)'
    __sTagSalesIncRate_y2_y1 ='매출증가율[-Y2][-Y1](%)'
    __sTagSalesIncRate_lqy1_lq ='매출증가율[Q-Y1][Q](%)'
    __sTagSalesIncRateAvg ='매출증가율[AVG](%)'
    __sTagSalesIncRateTrend_y3_y1 ='매출증가추세[-Y3][-Y1]'
    __sTagSalesIncRateTrend_y3_PredQ = '매출증가추세[-Y3][Q+]'

    # BsProf 성장지표
    __sTagBsProf_y3 = '영업이익[-Y3](억)'
    __sTagBsProf_y2 = '영업이익[-Y2](억)'
    __sTagBsProf_y1 = '영업이익[-Y1](억)'
    __sTagBsProfLastQ = '영업이익[Q](억)'
    __sTagBsProfLastQ_y1 = '영업이익[Q-Y1](억)'
    __sTagBsProfPredQ = '영업이익[Q+](억)'

    __sTagBsProfIncRate_y3_y2 = '영익증가율[-Y3][-Y2](%)'
    __sTagBsProfIncRate_y2_y1 = '영익증가율[-Y2][-Y1](%)'
    __sTagBsProfIncRate_lqy1_lq = '영익증가율[Q-Y1][Q](%)'
    __sTagBsProfIncRateAvg = '영익증가율[AVG](%)'
    __sTagBsProfIncRateTrend_y3_y1 = '영익증가추세[-Y3][-Y1]'
    __sTagBsProfIncRateTrend_y3_PredQ = '영익증가추세[-Y3][Q+]'

    # Group Column Names
    __sTagStockGrp = '종목'
    __sTagSrimExpGrp = '기대수익_SRIM'
    __sTagSrimPriceGrp = '적정주가_SRIM'
    __sTagKPriceGrp = '적정주가_KPRICE'
    __sTagGrowthGrp = '성장지표'
    __sTagSafeIndGrp = '안전지표'
    __sTagSubIndGrp = '보조지표'
    __sTagDebGrp = '부채항목'


class CStockResultModel:
    def __init__(self):
        self.__Data = CStockResultData()

    def Data(self):
        return self.__Data

    def AppendData(self, stockCode : str, fnData: CFnguideData , customData: CStockCustomData):
        self.__Data.AppendData(stockCode, fnData, customData)

    def ClearData(self):
        self.__Data.de

    def ExportExcel(self, sOutPath, sSheetName):
        resDf = self.__Data.ResultDf()

        if not os.path.exists(sOutPath):
            with pd.ExcelWriter(sOutPath, mode='w', engine='openpyxl') as writer:
                resDf.to_excel(writer, sheet_name=sSheetName, index=False)
                # writer.save()
        else:
            wb = load_workbook(sOutPath)
            if sSheetName in wb.sheetnames:
                with pd.ExcelWriter(sOutPath, mode = 'a', engine='openpyxl') as writer:
                    writer.book = wb
                    writer.sheets = dict( (ws.title,ws) for ws in wb.worksheets)
                    resDf.to_excel(writer, sheet_name=sSheetName, header=False,
                                   startrow=writer.sheets[sSheetName].max_row, index=False)
                    # writer.save()
            else:
                with pd.ExcelWriter(sOutPath, mode='a', engine='openpyxl') as writer:
                    resDf.to_excel(writer, sheet_name=sSheetName, index=False)

                    # writer.save()

    def ExportSummaryExcel(self, sOutPath, sSheetName):
        resDf = self.__Data.ResultDf()
        summaryDf = resDf[self.__Data.SummaryCols()]
        summaryDf.columns = self.__Data.SummaryGroupCols()

        if not os.path.exists(sOutPath):
            with pd.ExcelWriter(sOutPath, mode='w', engine='openpyxl') as writer:
                summaryDf.to_excel(writer, sheet_name=sSheetName)
                # 멀티 컬럼으로 EXCEL Export시 헤더 아래 Empty row 삭제
                sheet = writer.sheets[sSheetName]
                sheet.delete_rows(3)
        else:
            wb = load_workbook(sOutPath)
            if sSheetName in wb.sheetnames:
                with pd.ExcelWriter(sOutPath, mode='a', engine='openpyxl') as writer:
                    writer.book = wb
                    writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
                    # 멀티 컬럼 컬럼은 to_excel(index=False)를 지원하지 않는다
                    summaryDf.to_excel(writer, sheet_name=sSheetName, header=False, #index=False
                                       startrow=writer.sheets[sSheetName].max_row-1)
                    # writer.save()
            else:
                with pd.ExcelWriter(sOutPath, mode='a', engine='openpyxl') as writer:
                    summaryDf.to_excel(writer, sheet_name=sSheetName)
                    # 멀티 컬럼으로 EXCEL Export시 헤더 아래 Empty row 삭제
                    sheet = writer.sheets[sSheetName]
                    sheet.delete_rows(3)

        # 위에서 engine을 openpyxl로 하게되면 아래 코드가 호환되지 않는다.
        # 그러나 append 모드로 열기위해서는 위 엔진을 사용해야한다.

        # resDf = self.__Data.ResultDf()
        # summaryDf = resDf[self.__Data.SummaryCols()]
        # summaryDf.columns = self.__Data.SummaryGroupCols()
        #
        # writer = pd.ExcelWriter(sOutPath)
        # summaryDf.to_excel(writer,sheet_name=sSheetName)
        #
        # # 멀티 컬럼으로 EXCEL Export시 헤더 아래 empty row가 생기는 문제 처리
        # writer.sheets[sSheetName].set_row(2,None,None,{'hidden':True})
        # writer.save()





